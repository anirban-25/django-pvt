import sys
import os
import shutil
import time
import datetime
import requests
import uuid
import json
import traceback
import xlsxwriter as xlsxwriter
from openpyxl import load_workbook

from django.conf import settings
from django.core.management.base import BaseCommand

from api.models import DME_Files
from api.views_external_apis import do_bulk_pricing

if settings.ENV == "local":
    RESULT_DIR = "./static/uploaded/pricing_only/result/"
    SRC_DIR = "./static/uploaded/pricing_only/indata/"
    SRC_INPROGRESS_DIR = "./static/uploaded/pricing_only/inprogress/"
    SRC_ACHIEVE_DIR = "./static/uploaded/pricing_only/achieve/"
else:
    RESULT_DIR = "/var/www/html/dme_api/static/uploaded/pricing_only/result/"
    SRC_DIR = "/var/www/html/dme_api/static/uploaded/pricing_only/indata/"
    SRC_INPROGRESS_DIR = (
        "/var/www/html/dme_api/static/uploaded/pricing_only/inprogress/"
    )
    SRC_ACHIEVE_DIR = "/var/www/html/dme_api/static/uploaded/pricing_only/achieve/"


def replace_null(array):
    for item in array:
        for attr in item:
            if item and item[attr] == None:
                item[attr] = ""

    return array


def _update_file_info(fname, fpath, note):
    print(f"#800 - {fname}, {note}")
    modified_fpath = fpath.replace("../dme_api/", "")
    file = DME_Files.objects.get(file_name=fname)
    file.file_path = modified_fpath
    file.note = note
    file.save()


def _insert_file_info(fname, fpath, note):
    print(f"#801 - {fname}, {note}")
    modified_fpath = fpath.replace("../dme_api/", "")

    # file = DME_Files.objects.get(file_name=fname, file_type="pricing-result")
    # file.delete()

    file = DME_Files()
    file.file_name = fname
    file.file_path = modified_fpath
    file.z_createdByAccount = "management script"
    file.note = note
    file.file_type = "pricing-result"
    file.save()


def read_xls(file):
    wb = load_workbook(filename=file, data_only=True)

    if "Import Headers and Lines" in wb.sheetnames:
        worksheet0 = wb["Import Headers and Lines"]
    else:
        print("#910 - File format is not supported.")
        return

    bookings = []
    booking_lines = []

    row = 6
    last_pk_booking_id = None
    while True:
        pk_booking_id = worksheet0["A%i" % row].value

        if pk_booking_id == None:
            break
        else:
            pk_booking_id = str(pk_booking_id) + "_pricing_only"

        if int(worksheet0["CX%i" % row].value) == 0:
            print(f"@804 - {row} qty 0")
            row += 1
            continue

        if not last_pk_booking_id or pk_booking_id != last_pk_booking_id:
            last_pk_booking_id = pk_booking_id
            booking = {
                "pk_booking_id": pk_booking_id,
                "puPickUpAvailFrom_Date": str(
                    datetime.datetime.now() + datetime.timedelta(days=1)
                ),
                "b_clientReference_RA_Numbers": worksheet0["B%i" % row].value,
                "puCompany": worksheet0["K%i" % row].value
                if worksheet0["K%i" % row].value
                else "HARDCODED_00",
                "pu_Contact_F_L_Name": worksheet0["T%i" % row].value
                if worksheet0["T%i" % row].value
                else "HARDCODED_01",
                "pu_Email": worksheet0["W%i" % row].value
                if worksheet0["W%i" % row].value
                else "pu@email.com",
                "pu_Phone_Main": worksheet0["U%i" % row].value
                if worksheet0["U%i" % row].value
                else "419294339",
                "pu_Address_Street_1": worksheet0["M%i" % row].value,
                "pu_Address_street_2": worksheet0["N%i" % row].value,
                "pu_Address_Country": worksheet0["Q%i" % row].value,
                "pu_Address_PostalCode": str(worksheet0["R%i" % row].value),
                "pu_Address_State": worksheet0["P%i" % row].value,
                "pu_Address_Suburb": worksheet0["O%i" % row].value,
                "deToCompanyName": worksheet0["AK%i" % row].value
                if worksheet0["AK%i" % row].value
                else "HARDCODED_10",
                "de_to_Contact_F_LName": worksheet0["AT%i" % row].value
                if worksheet0["AT%i" % row].value
                else "HARDCODED_11",
                "de_Email": worksheet0["AW%i" % row].value
                if worksheet0["AW%i" % row].value
                else "de@email.com",
                "de_to_Phone_Main": worksheet0["AV%i" % row].value
                if worksheet0["AV%i" % row].value
                else "419294339",
                "de_To_Address_Street_1": worksheet0["AM%i" % row].value,
                "de_To_Address_Street_2": worksheet0["AN%i" % row].value,
                "de_To_Address_Country": worksheet0["AQ%i" % row].value,
                "de_To_Address_PostalCode": str(worksheet0["AR%i" % row].value),
                "de_To_Address_State": worksheet0["AP%i" % row].value,
                "de_To_Address_Suburb": worksheet0["AO%i" % row].value,
                "client_warehouse_code": worksheet0["I%i" % row].value,
                "pu_Address_Type": worksheet0["EG%i" % row].value or "business",
                "de_To_AddressType": worksheet0["EH%i" % row].value or "business",
                "b_booking_tail_lift_pickup": worksheet0["EI%i" % row].value or 0,
                "b_booking_tail_lift_deliver": worksheet0["EJ%i" % row].value or 0,
                "vx_serviceName": "R",  # hardcoded
                "b_client_name": "Pricing-Only",  # hardcoded
                "kf_client_id": "461162D2-90C7-BF4E-A905-0242ac130003",  # hardcoded - pricing_only
                "pu_PickUp_By_Date": str(
                    (datetime.datetime.now() + datetime.timedelta(days=1)).date()
                ),
                "pu_PickUp_By_Time_Hours": "10",
                "pu_PickUp_By_Time_Minutes": "00",
                "b_dateBookedDate": "",
                "b_client_name": "Pricing-Only",
            }
            bookings.append(booking)

        booking_line = {
            "pk_lines_id": str(uuid.uuid4()),
            "fk_booking_id": pk_booking_id,
            "e_dimWidth": worksheet0["DE%i" % row].value,
            "e_dimHeight": worksheet0["DF%i" % row].value,
            "e_dimLength": worksheet0["DD%i" % row].value,
            "e_dimUOM": worksheet0["DC%i" % row].value,
            "e_weightPerEach": worksheet0["DH%i" % row].value,
            "e_weightUOM": worksheet0["DG%i" % row].value,
            "e_item": worksheet0["CZ%i" % row].value,
            "packagingType": worksheet0["CW%i" % row].value,
            "e_type_of_packaging": worksheet0["CW%i" % row].value,
            "e_qty": worksheet0["CX%i" % row].value,
            "packed_status": "original",
        }
        booking_lines.append(booking_line)
        row += 1

    return bookings, booking_lines


def do_process(fpath, fname):
    # Create RESULT_DIR folder if not exist
    if not os.path.exists(RESULT_DIR):
        os.makedirs(RESULT_DIR)

    file_name_without_ext = fname.split(".")[0]
    file_name = f"{file_name_without_ext}_result_{str(datetime.datetime.now().strftime('%d-%m-%Y %H_%M_%S'))}.xlsx"
    file_path = f"{RESULT_DIR}/{file_name}"

    # Initialize XLSWriter
    workbook = xlsxwriter.Workbook(file_path, {"remove_timezone": True})
    worksheet = workbook.add_worksheet()
    bold = workbook.add_format({"bold": 1, "align": "left"})
    date_format = workbook.add_format({"num_format": "dd/mm/yyyy"})
    time_format = workbook.add_format({"num_format": "hh:mm:ss"})
    worksheet.set_column(0, 35, width=30)
    row = 2
    col = 0

    # Add header
    worksheet.write("A1", "pk_booking_id", bold)
    worksheet.write("B1", "puPickUpAvailFrom_Date", bold)
    worksheet.write("C1", "b_clientReference_RA_Numbers", bold)
    worksheet.write("D1", "puCompany", bold)
    worksheet.write("E1", "pu_Contact_F_L_Name", bold)
    worksheet.write("F1", "pu_Email", bold)
    worksheet.write("G1", "pu_Phone_Main", bold)
    worksheet.write("H1", "pu_Address_Street_1", bold)
    worksheet.write("I1", "pu_Address_Street_2", bold)
    worksheet.write("J1", "pu_Address_Country", bold)
    worksheet.write("K1", "pu_Address_PostalCode", bold)
    worksheet.write("L1", "pu_Address_State", bold)
    worksheet.write("M1", "pu_Address_Suburb", bold)
    worksheet.write("N1", "deToCompanyName", bold)
    worksheet.write("O1", "de_to_Contact_F_LName", bold)
    worksheet.write("P1", "de_Email", bold)
    worksheet.write("Q1", "de_to_Phone_Main", bold)
    worksheet.write("R1", "de_To_Address_Street_1", bold)
    worksheet.write("S1", "de_To_Address_Street_2", bold)
    worksheet.write("T1", "de_To_Address_Country", bold)
    worksheet.write("U1", "de_To_Address_PostalCode", bold)
    worksheet.write("V1", "de_To_Address_State", bold)
    worksheet.write("W1", "de_To_Address_Suburb", bold)
    worksheet.write("X1", "client_warehouse_code", bold)
    worksheet.write("Y1", "pu_Address_Type", bold)
    worksheet.write("Z1", "de_to_address_type", bold)
    worksheet.write("AA1", "Intentional Blank", bold)
    worksheet.write("AB1", "No", bold)
    worksheet.write("AC1", "Transporter", bold)
    worksheet.write("AD1", "Service (Vehicle)", bold)
    worksheet.write("AE1", "Transport Days (working)", bold)
    worksheet.write("AF1", "FP Cost (Ex GST)", bold)
    worksheet.write("AG1", "FP Extra`s (Ex GST)", bold)
    worksheet.write("AH1", "FP Fuel Levy %", bold)
    worksheet.write("AI1", "FP Fuel Levy Amount", bold)
    worksheet.write("AJ1", "FP Total Cost (Ex GST)", bold)
    worksheet.write("AK1", "DME Client Markup %", bold)
    worksheet.write("AL1", "Intentional Blank", bold)
    worksheet.write("AM1", "Cost $", bold)
    worksheet.write("AN1", "FP Fuel Levy %", bold)
    worksheet.write("AO1", "FP Fuel Levy Amount", bold)
    worksheet.write("AP1", "Extra $", bold)
    worksheet.write("AQ1", "Total $ (Ex. GST)", bold)

    bookings, booking_lines = read_xls(fpath)
    bookings = replace_null(bookings)
    booking_lines = replace_null(booking_lines)
    results = do_bulk_pricing(bookings, booking_lines)

    for result in results:
        booking = result["booking"]
        pricings = result["pricings"]

        for index, pricing in enumerate(pricings):
            worksheet.write(row, col + 0, booking["pk_booking_id"])
            worksheet.write(row, col + 1, booking["puPickUpAvailFrom_Date"])
            worksheet.write(row, col + 2, booking["b_clientReference_RA_Numbers"])
            worksheet.write(row, col + 3, booking["puCompany"])
            worksheet.write(row, col + 4, booking["pu_Contact_F_L_Name"])
            worksheet.write(row, col + 5, booking["pu_Email"])
            worksheet.write(row, col + 6, booking["pu_Phone_Main"])
            worksheet.write(row, col + 7, booking["pu_Address_Street_1"])
            worksheet.write(row, col + 8, booking["pu_Address_street_2"])
            worksheet.write(row, col + 9, booking["pu_Address_Country"])
            worksheet.write(row, col + 10, booking["pu_Address_PostalCode"])
            worksheet.write(row, col + 11, booking["pu_Address_State"])
            worksheet.write(row, col + 12, booking["pu_Address_Suburb"])
            worksheet.write(row, col + 13, booking["deToCompanyName"])
            worksheet.write(row, col + 14, booking["de_to_Contact_F_LName"])
            worksheet.write(row, col + 15, booking["de_Email"])
            worksheet.write(row, col + 16, booking["de_to_Phone_Main"])
            worksheet.write(row, col + 17, booking["de_To_Address_Street_1"])
            worksheet.write(row, col + 18, booking["de_To_Address_Street_2"])
            worksheet.write(row, col + 19, booking["de_To_Address_Country"])
            worksheet.write(row, col + 20, booking["de_To_Address_PostalCode"])
            worksheet.write(row, col + 21, booking["de_To_Address_State"])
            worksheet.write(row, col + 22, booking["de_To_Address_Suburb"])
            worksheet.write(row, col + 23, booking["client_warehouse_code"])
            worksheet.write(row, col + 24, booking["pu_Address_Type"])
            worksheet.write(row, col + 25, booking["de_To_AddressType"])
            worksheet.write(row, col + 26, "")
            worksheet.write(row, col + 27, index + 1)
            worksheet.write(
                row,
                col + 28,
                f"{pricing['freight_provider']} ({pricing['account_code']})",
            )
            worksheet.write(row, col + 29, pricing["service_name"])
            worksheet.write(row, col + 30, pricing["etd"])
            worksheet.write(row, col + 31, round(float(pricing["fee"]), 2))
            surcharge_total = (
                pricing["surcharge_total"] if pricing["surcharge_total"] else 0
            )
            worksheet.write(row, col + 32, round(float(surcharge_total), 2))
            worksheet.write(
                row, col + 33, float(pricing["mu_percentage_fuel_levy"]) * 100
            )
            worksheet.write(row, col + 34, round(float(pricing["fuel_levy_base"]), 2))
            worksheet.write(
                row,
                col + 35,
                round(
                    float(pricing["fee"])
                    + float(pricing["fuel_levy_base"])
                    + float(surcharge_total),
                    2,
                ),
            )
            worksheet.write(row, col + 36, pricing["client_mark_up_percent"] * 100)
            worksheet.write(row, col + 37, "")
            worksheet.write(row, col + 38, round(float(pricing["cost_dollar"]), 2))
            worksheet.write(
                row,
                col + 39,
                round(float(pricing["mu_percentage_fuel_levy"]) * 100, 2),
            )
            worksheet.write(
                row, col + 40, round(float(pricing["fuel_levy_base_cl"]), 2)
            )
            worksheet.write(row, col + 41, pricing["surcharge_total_cl"])
            worksheet.write(
                row,
                col + 42,
                round(
                    float(pricing["surcharge_total_cl"])
                    if pricing["surcharge_total_cl"]
                    else 0,
                    2,
                ),
            )
            worksheet.write(row, col + 43, pricing["client_mu_1_minimum_values"])
            worksheet.write(row, col + 44, pricing["surcharge_total_cl"])

            row += 1

    workbook.close()
    return file_path


class Command(BaseCommand):
    def handle(self, *args, **options):
        print("----- Bulk Pricing-Only ... -----")
        print("#900 Started %s" % datetime.datetime.now())

        for fname in os.listdir(SRC_DIR):
            fpath = os.path.join(SRC_DIR, fname)

            if os.path.isfile(fpath) and fname.endswith(".xlsx"):
                try:
                    shutil.move(SRC_DIR + fname, SRC_INPROGRESS_DIR + fname)
                    _update_file_info(
                        fname,
                        SRC_INPROGRESS_DIR + fname,
                        "In progress: 0%",
                    )
                    file_path = do_process(SRC_INPROGRESS_DIR + fname, fname)
                    shutil.move(SRC_INPROGRESS_DIR + fname, SRC_ACHIEVE_DIR + fname)

                    # TEST Usage
                    # shutil.move(SRC_ACHIEVE_DIR + fname, SRC_DIR + fname)

                    _update_file_info(fname, SRC_ACHIEVE_DIR + fname, "Done: 100%")
                    file_name = file_path.split("/")[-1]
                    _insert_file_info(file_name, file_path, "Generated")
                except Exception as e:
                    traceback.print_exc()
                    _update_file_info(
                        fname,
                        SRC_INPROGRESS_DIR + fname,
                        f"Stopped... {str(e)}",
                    )

        print("#999 Finished %s\n\n\n" % datetime.datetime.now())
