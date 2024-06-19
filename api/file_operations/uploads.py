import os
from datetime import datetime
from django.conf import settings
import openpyxl

from api.common import trace_error
from api.models import (
    Bookings,
    DME_employees,
    DME_clients,
    Client_employees,
    Client_Products,
    Fp_freight_providers,
    DME_Files,
    Dme_attachments,
)
from django.forms.models import model_to_dict
from api.utils import (
    clearFileCheckHistory,
    getFileCheckHistory,
    save2Redis,
)


def get_upload_status(request):
    return JsonResponse({"status_code": 0})

    # result = getFileCheckHistory(request.GET.get("filename"))
    #
    # if result == 0:
    #     return JsonResponse({"status_code": 0})
    # elif result == "success":
    #     return JsonResponse({"status_code": 1})
    # else:
    #     return JsonResponse({"status_code": 2, "errors": result})


def _save_import_file(dme_account_num, file, client_company_name):
    if settings.ENV in ["prod", "dev"]:  # PROD & DEV
        file_path = "/var/www/html/dme_api/media/onedrive"
        file_name = f"{str(dme_account_num)}_{file.name}"
    else:  # LOCAL
        file_path = "./static/uploaded"
        file_name = file.name

    if not os.path.isdir(file_path):
        os.makedirs(file_path)

    full_path = f"{file_path}/{file_name}"

    with open(full_path, "wb+") as destination:
        for chunk in file.chunks():
            destination.write(chunk)

    # clearFileCheckHistory(f"str(dme_account_num)_{file.name}")


def upload_import_file(user_id, file, uploader):
    # Plum
    if uploader == "Plum Products Australia Ltd":
        if settings.ENV in ["prod", "dev"]:  # PROD & DEV
            file_path = "/dme_sftp/sapb1/order_transaction_csvs/indata"
        else:  # LOCAL
            file_path = "./static/uploaded"

        file_name = file.name
        full_path = f"{file_path}/{file_name}"
        with open(full_path, "wb+") as destination:
            for chunk in file.chunks():
                destination.write(chunk)
        return file_name

    dme_employee = DME_employees.objects.filter(fk_id_user=user_id).first()
    user_type = "DME" if dme_employee else "CLIENT"

    if user_type == "DME":
        dme_account_num = DME_clients.objects.get(company_name=uploader).dme_account_num
        client_company_name = "DME"
    else:
        client_employee = Client_employees.objects.get(fk_id_user=int(user_id))
        dme_account_num = client_employee.fk_id_dme_client.dme_account_num
        client_company_name = DME_clients.objects.get(
            pk_id_dme_client=client_employee.fk_id_dme_client_id
        ).company_name

    file_name = f"{str(dme_account_num)}_{file.name}"

    save2Redis(file_name + "_l_000_client_acct_number", dme_account_num)
    _save_import_file(dme_account_num, file, client_company_name)
    return file_name


def upload_attachment_file(user_id, file, booking_id, upload_option):
    try:
        try:
            client = DME_clients.objects.get(pk_id_dme_client=user_id)
        except DME_clients.DoesNotExist as e:
            client = DME_clients.objects.get(company_name="Pricing-Only")

        booking = Bookings.objects.get(id=booking_id)
        fp = Fp_freight_providers.objects.get(
            fp_company_name=booking.vx_freight_provider
        )
        name, extension = os.path.splitext(file.name)

        if upload_option == "attachment":
            fp_dir_name = (
                f"{fp.fp_company_name.lower()}_{fp.fp_address_country.lower()}"
            )
            file_path = f"{settings.STATIC_PUBLIC}/attachments/{fp_dir_name}"

            if not os.path.isdir(file_path):
                os.makedirs(file_path)

            file_name = (
                f"{name}-{str(datetime.now().strftime('%Y%m%d_%H%M%S'))}{extension}"
            )
            full_path = f"{file_path}/{file_name}"
        elif upload_option in ["label", "pod"]:
            fp_dir_name = (
                f"{fp.fp_company_name.lower()}_{fp.fp_address_country.lower()}"
            )

            if upload_option == "label":
                file_path = f"{settings.STATIC_PUBLIC}/pdfs/{fp_dir_name}"
            else:
                file_path = f"{settings.STATIC_PUBLIC}/imgs/{fp_dir_name}"

            if not os.path.isdir(file_path):
                os.makedirs(file_path)

            if upload_option == "label":
                file_name = f"DME{str(booking.b_bookingID_Visual)}{extension}"
                booking.z_label_url = f"{fp.fp_company_name.lower()}_{fp.fp_address_country.lower()}/{file_name}"
            elif upload_option == "pod" and not "sog" in name.lower():
                file_name = f"POD_DME{str(booking.b_bookingID_Visual)}{extension}"
                booking.z_pod_url = f"{fp.fp_company_name.lower()}_{fp.fp_address_country.lower()}/{file_name}"
            elif upload_option == "pod" and "sog" in name.lower():
                file_name = f"POD_SOG_DME{str(booking.b_bookingID_Visual)}{extension}"
                booking.z_pod_signed_url = f"{fp.fp_company_name.lower()}_{fp.fp_address_country.lower()}/{file_name}"

            full_path = f"{file_path}/{file_name}"
            booking.z_ModifiedTimestamp = datetime.now()
            booking.save()

        with open(full_path, "wb+") as destination:
            for chunk in file.chunks():
                destination.write(chunk)

        if upload_option == "attachment":
            dme_attachment = Dme_attachments(
                fk_id_dme_client=client,
                fk_id_dme_booking=booking.pk_booking_id,
                fileName=full_path,
                linkurl="22",
                upload_Date=datetime.now(),
            )
            dme_attachment.save()

        return {
            "status": "success",
            "file_path": f"{fp_dir_name}/{file_name}",
            "type": upload_option,
        }
    except Exception as e:
        trace_error.print()
        return {
            "status": "failed",
            "type": upload_option,
        }


def upload_pricing_only_file(user_id, username, file, upload_option):
    dme_file = DME_Files.objects.create(
        file_name=f"__{file.name}",
        z_createdByAccount=username,
        file_type="pricing-only",
        file_extension="xlsx",
        note="Uploaded to get Pricings only",
    )
    dir_path = f"./static/uploaded/pricing_only/indata/"
    dme_file.file_name = f"{dme_file.pk}__{file.name}"
    dme_file.file_path = (
        f"./static/uploaded/pricing_only/indata/{dme_file.pk}__{file.name}"
    )
    dme_file.save()

    if not os.path.isdir(dir_path):
        os.makedirs(dir_path)

    with open(dme_file.file_path, "wb+") as destination:
        for chunk in file.chunks():
            destination.write(chunk)

    return {
        "status": "success",
        "file_name": dme_file.file_name,
        "type": upload_option,
    }


def upload_pricing_rule_file(user_id, username, file, upload_option, rule_type):
    dme_file = DME_Files.objects.create(
        file_name=f"__{file.name}",
        z_createdByAccount=username,
        file_type="pricing-rule",
        file_extension="xlsx",
        note="Uploaded to import Pricings Rules sheet",
    )
    dir_path = f"./static/uploaded/pricing_rule/indata/"
    full_path = (
        f"./static/uploaded/pricing_rule/indata/{dme_file.pk}__{rule_type}__{file.name}"
    )
    dme_file.file_name = f"{dme_file.pk}__{rule_type}__{file.name}"
    dme_file.file_path = full_path
    dme_file.save()

    if not os.path.isdir(dir_path):
        os.makedirs(dir_path)

    with open(full_path, "wb+") as destination:
        for chunk in file.chunks():
            destination.write(chunk)

    return {
        "status": "success",
        "file_name": dme_file.file_name,
        "type": upload_option,
    }


def upload_client_products_file(user_id, username, client_id, file):
    wb = openpyxl.load_workbook(file)
    ws = wb["Product Import"]

    if client_id is None:
        client_employee = Client_employees.objects.get(fk_id_user=int(user_id))
        dme_client = DME_clients.objects.get(
            pk_id_dme_client=client_employee.fk_id_dme_client_id
        )
    else:
        dme_client = DME_clients.objects.get(pk_id_dme_client=client_id)

    def check_data(data):
        not_empty_cols = [
            "parent_model_number",
            "child_model_number",
            "description",
            "qty",
        ]
        for key in data:
            if key in not_empty_cols and (data[key] == None or data[key] == "NULL"):
                return False
            elif key not in not_empty_cols and data[key] == "NULL":
                data[key] = None
        return data

    import_success_results = []
    empty_field_rows = []
    wrong_type_rows = []
    try:
        Client_Products.objects.filter(fk_id_dme_client=dme_client).delete()
        delete_status = "success"
    except Exception as e:
        delete_status = "failed"

    success_count = 0
    failure_count = 0
    created_products = []
    for r in range(2, ws.max_row + 1):
        data = {
            "fk_id_dme_client": dme_client,
            "parent_model_number": ws.cell(row=r, column=1).value,
            "child_model_number": ws.cell(row=r, column=2).value,
            "description": ws.cell(row=r, column=3).value,
            "qty": ws.cell(row=r, column=4).value,
            "e_dimUOM": str(ws.cell(row=r, column=5).value).lower(),
            "e_weightUOM": str(ws.cell(row=r, column=6).value).lower(),
            "e_dimLength": ws.cell(row=r, column=7).value,
            "e_dimWidth": ws.cell(row=r, column=8).value,
            "e_dimHeight": ws.cell(row=r, column=9).value,
            "e_weightPerEach": ws.cell(row=r, column=10).value,
            "z_createdByAccount": username,
        }

        valid_data = check_data(data)
        if valid_data:
            try:
                created = Client_Products.objects.create(**valid_data)
                created.save()
                success_count = success_count + 1
                created_products.append(model_to_dict(created))
                import_success_results.append(r)
            except Exception as e:
                failure_count = failure_count + 1
                wrong_type_rows.append(r)
        else:
            failure_count = failure_count + 1
            empty_field_rows.append(r)

    return {
        "file_name": file.name,
        "delete_status": delete_status,
        "import_status": {
            "success_count": success_count,
            "failure_count": failure_count,
            "success_rows": import_success_results,
            "failure_rows": {
                "empty_field_error": empty_field_rows,
                "wrong_type_error": wrong_type_rows,
            },
        },
        "created_products": created_products,
    }
